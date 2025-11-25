#!/usr/bin/env .venv/bin/python3
"""
Import employees from XLSX file directly to DynamoDB
November 2025 - Latest Version
"""
import boto3
from openpyxl import load_workbook
from datetime import datetime
from decimal import Decimal

# Excel file path
XLSX_FILE = '/Users/robwinters/Downloads/Active_Headcount__location_and_email_ (7).xlsx'

# Initialize DynamoDB
dynamodb = boto3.resource('dynamodb', region_name='us-east-2')
table = dynamodb.Table('panda-employees')

def get_all_existing_employees():
    """Get all existing employee IDs from DynamoDB"""
    print("Fetching existing employees from DynamoDB...")
    response = table.scan(ProjectionExpression='id')
    employee_ids = set()

    for item in response.get('Items', []):
        if 'id' in item:
            employee_ids.add(item['id'])

    # Handle pagination
    while 'LastEvaluatedKey' in response:
        response = table.scan(
            ProjectionExpression='id',
            ExclusiveStartKey=response['LastEvaluatedKey']
        )
        for item in response.get('Items', []):
            if 'id' in item:
                employee_ids.add(item['id'])

    print(f"Found {len(employee_ids)} existing employees in DynamoDB")
    return employee_ids

def parse_supervisor_name(supervisor_str):
    """
    Parse supervisor name from 'Last, First' format to 'First Last'
    """
    if not supervisor_str or str(supervisor_str).strip() == '':
        return ''

    supervisor_str = str(supervisor_str).strip()

    # Handle 'Last, First' format
    if ',' in supervisor_str:
        parts = supervisor_str.split(',')
        if len(parts) == 2:
            last_name = parts[0].strip()
            first_name = parts[1].strip()
            return f"{first_name} {last_name}"

    return supervisor_str

def import_employees():
    """Import employees from XLSX to DynamoDB"""

    # Get all existing employee IDs
    existing_employee_ids = get_all_existing_employees()

    updated_count = 0
    new_count = 0
    with_emails = 0
    without_emails = 0
    terminated_count = 0
    csv_employee_ids = set()

    print(f"\nImporting from: {XLSX_FILE}\n")

    # Load workbook
    wb = load_workbook(XLSX_FILE, data_only=True)
    ws = wb.active

    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    print(f"Headers found: {headers}")
    print(f"Total rows: {ws.max_row}")
    print()

    # Process each row
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Create dictionary from row
        row_dict = dict(zip(headers, row))

        # Skip empty rows
        if not row_dict.get('Employee Id') or not row_dict.get('Last Name'):
            continue

        employee_id = str(row_dict.get('Employee Id', '')).strip()
        first_name = str(row_dict.get('First Name', '')).strip()
        last_name = str(row_dict.get('Last Name', '')).strip()
        email = str(row_dict.get('Work Email', '') or '').strip().lower()
        phone = str(row_dict.get('Phone', '') or '').strip()
        hire_date = row_dict.get('Employment Date', '') or row_dict.get('Hire Date', '')
        termination_date = row_dict.get('Termination Date', '') or ''
        position = str(row_dict.get('Position', '') or row_dict.get('Job Title', '') or '').strip()
        department = str(row_dict.get('Department', '') or '').strip()
        office = str(row_dict.get('Current Work Location Name', '') or row_dict.get('Office', '') or '').strip()
        supervisor_raw = row_dict.get('Supervisor', '') or row_dict.get('Manager', '') or ''
        supervisor = parse_supervisor_name(supervisor_raw)

        # Convert hire date if it's a datetime object
        if isinstance(hire_date, datetime):
            hire_date = hire_date.strftime('%Y-%m-%d')
        elif hire_date:
            hire_date = str(hire_date)
        else:
            hire_date = ''

        # Convert termination date if it's a datetime object
        if isinstance(termination_date, datetime):
            termination_date = termination_date.strftime('%Y-%m-%d')
        elif termination_date:
            termination_date = str(termination_date).strip()
        else:
            termination_date = ''

        # Determine if terminated
        is_terminated = bool(termination_date) or row_dict.get('Employee Status Code') == 'T'

        # Track email status
        if email:
            with_emails += 1
        else:
            without_emails += 1

        csv_employee_ids.add(employee_id)

        # Create employee record
        employee = {
            'id': employee_id,
            'employee_id': employee_id,
            'Employee Id': employee_id,
            'First Name': first_name,
            'Last Name': last_name,
            'Email': email,
            'Work Email': email,
            'Phone': phone,
            'Position': position,
            'Department': department,
            'office': office,
            'Current Work Location Name': office,
            'Employment Date': hire_date,
            'hire_date': hire_date,
            'supervisor': supervisor,
            'manager': supervisor,
            'points': Decimal('0'),
            'Panda Points': Decimal('0'),
            'password': 'Panda2025!',
            'points_manager': 'No',
            'points_budget': Decimal('0'),
            'merchandise_value': Decimal('0'),
            'Merchandise Value': Decimal('0'),
            'Merch Requested': '',
            'Merch Sent': 'No',
            'Merch Sent Date': '',
            'updated_at': datetime.now().isoformat()
        }

        # Handle termination
        if termination_date:
            employee['termination_date'] = termination_date
            employee['Termination Date'] = termination_date
            employee['Terminated'] = 'Yes'
            employee['status'] = 'inactive'
            employee['is_active'] = False
            terminated_count += 1
        else:
            employee['Terminated'] = 'No'
            employee['status'] = 'active'
            employee['is_active'] = True

        # Put item in DynamoDB
        try:
            table.put_item(Item=employee)

            if employee_id in existing_employee_ids:
                updated_count += 1
                status = "UPDATED"
            else:
                new_count += 1
                status = "NEW"

            term_status = " [TERMINATED]" if is_terminated else ""
            print(f"✓ {status}: {first_name} {last_name} (ID: {employee_id}){term_status}")

        except Exception as e:
            print(f"✗ ERROR importing {first_name} {last_name} (ID: {employee_id}): {str(e)}")

    # Summary
    print(f"\n{'='*60}")
    print(f"IMPORT COMPLETE")
    print(f"{'='*60}")
    print(f"New employees added: {new_count}")
    print(f"Existing employees updated: {updated_count}")
    print(f"Total employees in CSV: {len(csv_employee_ids)}")
    print(f"Employees with email: {with_emails}")
    print(f"Employees without email: {without_emails}")
    print(f"Terminated employees: {terminated_count}")
    print(f"{'='*60}\n")

if __name__ == '__main__':
    import_employees()
